from uuid import uuid4
import psycopg2.extras
from openpyxl import load_workbook

from config import get_pg_connection, SPREADSHEET_PATH


wb = load_workbook(SPREADSHEET_PATH)
skills = wb['Навыки']

psycopg2.extras.register_uuid()

for row in skills.iter_rows(values_only=True):
    if row[0]:
        query = """ INSERT INTO questions_skills (id, title) VALUES (%s, %s); """
        with get_pg_connection() as conn, conn.cursor() as cur:
            cur.execute(query, (uuid4(), row[0],))
    else:
        continue
